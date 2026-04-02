"""
FAQ・用語集JSONをExcelに書き出すスクリプト。
git post-commit hook または手動で実行する。

Usage:
    python scripts/export_to_excel.py
    python scripts/export_to_excel.py --output path/to/output.xlsx
"""
import json
import argparse
import sys
import io
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl が見つかりません: pip install openpyxl")
    sys.exit(1)

# ---- パス設定 ----
REPO_ROOT = Path(__file__).parent.parent
FAQ_JSON = REPO_ROOT / "data" / "faq" / "faq_master.json"
GLOSSARY_JSON = REPO_ROOT / "data" / "glossary" / "glossary_master.json"
OUTPUT_DIR = REPO_ROOT / "data" / "excel"

# ---- スタイル定数 ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # ネイビー
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")   # 薄青（偶数行）
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="C0C0C0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, headers: list[str], col_widths: list[int]):
    """1行目をヘッダースタイルに設定"""
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, values: list, alt: bool):
    fill = ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def build_faq_sheet(wb: openpyxl.Workbook):
    with open(FAQ_JSON, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("items", [])

    ws = wb.create_sheet("FAQ")
    headers = ["ID", "質問", "回答", "タグ", "関連FAQ", "ソース", "embedding_text"]
    widths  = [10,   40,    60,    25,    15,      20,     50]
    _style_header(ws, headers, widths)

    for i, item in enumerate(items, 2):
        tags = "、".join(item.get("tags") or [])
        related = "、".join(item.get("related_faq_ids") or [])
        values = [
            item.get("id", ""),
            item.get("question", ""),
            item.get("answer", ""),
            tags,
            related,
            item.get("source") or "",
            item.get("embedding_text", ""),
        ]
        _write_row(ws, i, values, alt=(i % 2 == 0))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    print(f"  FAQ シート: {len(items)} 件")


def build_glossary_sheet(wb: openpyxl.Workbook):
    with open(GLOSSARY_JSON, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("items", [])

    ws = wb.create_sheet("用語集")
    headers = ["ID", "種別", "用語", "表記ゆれ", "定義", "オペレーター向け定義",
               "関連用語", "関連FAQ", "カテゴリ", "タグ", "embedding_text"]
    widths  = [12,   10,    20,     25,       50,    50,
               25,    15,    15,     20,       50]
    _style_header(ws, headers, widths)

    for i, item in enumerate(items, 2):
        variants = "、".join(item.get("term_variants") or [])
        related_terms = "、".join(item.get("related_terms") or [])
        related_faqs = "、".join(item.get("related_faq_ids") or [])
        tags = "、".join(item.get("tags") or [])
        values = [
            item.get("id", ""),
            item.get("type", ""),
            item.get("term", ""),
            variants,
            item.get("definition", ""),
            item.get("definition_for_operator", ""),
            related_terms,
            related_faqs,
            item.get("category", ""),
            tags,
            item.get("embedding_text", ""),
        ]
        _write_row(ws, i, values, alt=(i % 2 == 0))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    print(f"  用語集シート: {len(items)} 件")


def build_summary_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("サマリー", 0)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = [
        ("エクスポート日時", now),
        ("FAQデータ", str(FAQ_JSON)),
        ("用語集データ", str(GLOSSARY_JSON)),
        ("", ""),
        ("シート構成", ""),
        ("FAQ", "FAQナレッジ一覧（全項目）"),
        ("用語集", "用語・定義一覧（全項目）"),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", help="出力ファイルパス（省略時は data/excel/ に日付付きで保存）")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if args.output:
        out_path = Path(args.output)
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = OUTPUT_DIR / f"srcc_knowledge_{stamp}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    print("Exporting to Excel...")
    build_summary_sheet(wb)
    build_faq_sheet(wb)
    build_glossary_sheet(wb)

    wb.save(out_path)
    print(f"  -> {out_path}")
    print("Done.")


if __name__ == "__main__":
    main()
